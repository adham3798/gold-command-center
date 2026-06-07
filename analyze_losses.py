# -*- coding: utf-8 -*-
"""Loss analysis: reconstruct the Nature-Cycle model per day, tag attributes,
find what the LOSING days have in common, write a formatted Excel workbook."""
import pandas as pd, urllib.parse, statistics as st
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SID='12ynlr46bvHSJLnLGs5Z1SrhhlCj6_w7qO6YHMDBY7gs'
def tab(t): return pd.read_csv('https://docs.google.com/spreadsheets/d/%s/gviz/tq?tqx=out:csv&sheet=%s'%(SID,urllib.parse.quote(t)))
def col(df,*n):
    import re; norm={re.sub(r'\s+',' ',str(c)).strip().lower():c for c in df.columns}
    for x in n:
        k=re.sub(r'\s+',' ',x).strip().lower()
        if k in norm:return norm[k]
    for x in n:
        for k,o in norm.items():
            if x.strip().lower() in k:return o
    return None
def clean(s):
    import re; s=str(s); m=re.search(r'\(([^)]+)\)',s); return m.group(1).strip() if m else s.strip()
def droot(d):
    d=int(d)
    while d>9:d=sum(int(c) for c in str(d))
    return d

RETRO=[('Mercury','2/26/2026','3/20/2026'),('Pluto','5/6/2026','10/16/2026'),('Mercury','6/29/2026','7/23/2026'),
       ('Neptune','7/7/2026','12/12/2026'),('Saturn','7/26/2026','12/10/2026'),('Mercury','10/24/2026','11/13/2026'),
       ('Venus','10/3/2026','11/14/2026'),('Uranus','9/10/2026','2/4/2026')]
def active_retros(ds):
    dt=datetime.strptime(ds,'%Y-%m-%d'); out=[]
    for p,s,e in RETRO:
        sd=datetime.strptime(s,'%m/%d/%Y'); ed=datetime.strptime(e,'%m/%d/%Y')
        if sd<=dt<=ed: out.append(p)
    return out

# load
gp=tab('gold_price'); gp.columns=[str(c).strip() for c in gp.columns]
gp['Date']=pd.to_datetime(gp['Date'],errors='coerce'); gp=gp.dropna(subset=['Close']).sort_values('Date')
P={}
for _,r in gp.iterrows():
    P[r['Date'].strftime('%Y-%m-%d')]={'o':float(r['Open']),'h':float(r['High']),'l':float(r['Low']),'c':float(r['Close']),'dir':str(r['Direction']).strip().upper()}
mr=tab('MOON_REAL'); mr.columns=[str(c).strip() for c in mr.columns]
cD,cS,cSt,cPh=col(mr,'Real Date'),col(mr,'Clean Moon Sign','Moon Sign'),col(mr,'Cycle Stage'),col(mr,'Moon Phase (Lunar Phase)','Moon Phase')
cDn=col(mr,'day number','Day Number')
mr[cD]=pd.to_datetime(mr[cD],errors='coerce'); mr=mr.dropna(subset=[cD])
MOON={r[cD].strftime('%Y-%m-%d'):{'sign':clean(r[cS]),'stage':str(r[cSt]).strip(),'phase':str(r[cPh]).strip(),'dn':r[cDn]} for _,r in mr.iterrows()}
sl=tab('SIGN_LIBRARY'); sl.columns=[str(c).strip() for c in sl.columns]
NAT={str(r['Sign']).strip():str(r['Nature']).strip().upper() for _,r in sl.iterrows() if str(r['Sign']).strip() and str(r['Sign']).strip().lower()!='nan'}

DAYS=sorted(d for d in P if d in MOON)
def matching(ds,sign,stage):
    ups,dns=[],[]
    for d in DAYS:
        if d>=ds:break
        if MOON[d]['sign']==sign and MOON[d]['stage']==stage:
            ch=P[d]['c']-P[d]['o']; (ups if ch>=0 else dns).append(abs(ch))
    gu=st.mean([abs(P[d]['c']-P[d]['o']) for d in DAYS if P[d]['c']>=P[d]['o']])
    gd=st.mean([abs(P[d]['c']-P[d]['o']) for d in DAYS if P[d]['c']<P[d]['o']])
    return (st.mean(ups) if len(ups)>=2 else gu),(st.mean(dns) if len(dns)>=2 else gd)

def model(ds,prev):
    pa=1 if P[prev]['dir']=='BULL' else -1
    nat=NAT.get(MOON[ds]['sign'],''); stage=MOON[ds]['stage']
    nine=droot(ds[-2:])==9; pnine=droot(prev[-2:])==9
    if pnine:   return -pa,'day after 9-cycle -> turn'
    if nine:    return pa,'9-cycle exhaustion -> push'
    if nat=='MOVABLE':
        same=MOON[prev]['sign']==MOON[ds]['sign']
        return (-pa,'movable 2nd-day -> pullback') if same else (pa,'movable 1st-day -> continue')
    if nat=='FIXED': return pa,'fixed -> continue trend'
    return (-pa,'finisher FINISH -> turn') if stage=='FINISH' else (pa,'finisher -> continue')

rows=[]
WD=['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
for i in range(1,len(DAYS)):
    ds,prev=DAYS[i],DAYS[i-1]; p=P[ds]; m=MOON[ds]
    d,rule=model(ds,prev); up,dn=matching(ds,m['sign'],m['stage']); mv=up if d>0 else dn
    o,h,l=p['o'],p['h'],p['l']
    if d>0: tp,slv=o+0.5*mv,o-mv; th,sh=h>=tp,l<=slv
    else:   tp,slv=o-0.5*mv,o+mv; th,sh=l<=tp,h>=slv
    win=th and not sh
    if win: reason='WIN'
    elif not th and not sh: reason='TP not reached (range too small)'
    elif th and sh: reason='whipsaw (TP & stop both hit)'
    else: reason='ran against (stop hit, no TP)'
    rows.append({'Date':ds,'Weekday':WD[datetime.strptime(ds,'%Y-%m-%d').weekday()],
        'Result':'WIN' if win else 'LOSS','Loss Reason':'' if win else reason,
        'Model Dir':'BUY' if d>0 else 'SELL','Rule (why)':rule,'Prior Day':P[prev]['dir'],
        'Moon Sign':m['sign'],'Nature':NAT.get(m['sign'],''),'Stage':m['stage'],
        'Day#':int(m['dn']) if pd.notna(m['dn']) else '','Date#':droot(ds[-2:]),
        'Important(3/7/9)':'Y' if droot(ds[-2:]) in (3,7,9) else '','9-date':'Y' if droot(ds[-2:])==9 else '',
        'Moon Phase':m['phase'],'Open':round(o,2),'High':round(h,2),'Low':round(l,2),'Close':round(p['c'],2),
        'Actual':p['dir'],'Range':round(h-l,2),'Exp Move':round(mv,2),'TP1':round(tp,2),'Stop':round(slv,2),
        'TP hit':'Y' if th else 'N','Stop hit':'Y' if sh else 'N','Retrogrades':', '.join(active_retros(ds))})

df=pd.DataFrame(rows)
loss=df[df['Result']=='LOSS'].copy()
W,L=len(df[df['Result']=='WIN']),len(loss)
print('Total %d | WIN %d | LOSS %d | win rate %.1f%%'%(len(df),W,L,W/len(df)*100))

def rate(field):
    g=df.groupby(field)['Result'].apply(lambda s:(s=='LOSS').mean()*100)
    n=df.groupby(field)['Result'].count()
    return pd.DataFrame({'Loss %':g.round(1),'Days':n}).sort_values('Loss %',ascending=False)

pat={}
for f in ['Nature','Stage','Moon Sign','Rule (why)','Weekday','Moon Phase','Important(3/7/9)','9-date','Prior Day']:
    pat[f]=rate(f)
lossreasons=loss['Loss Reason'].value_counts().rename_axis('Loss Reason').reset_index(name='Count')

# ── write workbook ──
wb=Workbook(); FNT='Arial'
hdr_fill=PatternFill('solid',start_color='1A2A40'); hdr_font=Font(name=FNT,bold=True,color='FFFFFF',size=10)
loss_fill=PatternFill('solid',start_color='FBE3E6'); win_fill=PatternFill('solid',start_color='E3F6EC')
thin=Side(style='thin',color='D0D0D0'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def style_header(ws,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(1,c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
    ws.freeze_panes='A2'
def write_df(ws,d,color_result=False):
    ws.append(list(d.columns))
    for _,r in d.iterrows(): ws.append(list(r.values))
    style_header(ws,len(d.columns))
    rescol=list(d.columns).index('Result')+1 if 'Result' in d.columns else None
    for rr in range(2,len(d)+2):
        for cc in range(1,len(d.columns)+1):
            cell=ws.cell(rr,cc); cell.font=Font(name=FNT,size=9); cell.border=border
        if color_result and rescol:
            v=ws.cell(rr,rescol).value
            fill=loss_fill if v=='LOSS' else win_fill
            for cc in range(1,len(d.columns)+1): ws.cell(rr,cc).fill=fill
    for cc,colname in enumerate(d.columns,1):
        w=max(len(str(colname)),*(len(str(x)) for x in d[colname].astype(str)))+2
        ws.column_dimensions[ws.cell(1,cc).column_letter].width=min(w,40)

ws=wb.active; ws.title='Losing Days'; write_df(ws,loss,color_result=True)
write_df(wb.create_sheet('All Days'),df,color_result=True)

# Patterns sheet
wp=wb.create_sheet('Loss Patterns')
wp.append(['LOSS RATE BY FACTOR  (higher = loses more often)']); wp['A1'].font=Font(name=FNT,bold=True,size=12)
row=3
for f,tblrows in pat.items():
    wp.cell(row,1,f).font=Font(name=FNT,bold=True,color='1A2A40',size=11); row+=1
    wp.cell(row,1,f); wp.cell(row,2,'Loss %'); wp.cell(row,3,'Days')
    for c in range(1,4): wp.cell(row,c).fill=hdr_fill; wp.cell(row,c).font=hdr_font; wp.cell(row,c).border=border
    row+=1
    for idx,r in tblrows.iterrows():
        wp.cell(row,1,str(idx)); wp.cell(row,2,r['Loss %']); wp.cell(row,3,int(r['Days']))
        for c in range(1,4): wp.cell(row,c).font=Font(name=FNT,size=9); wp.cell(row,c).border=border
        row+=1
    row+=1
wp.column_dimensions['A'].width=34; wp.column_dimensions['B'].width=10; wp.column_dimensions['C'].width=8

# Loss reasons + Why summary
wr=wb.create_sheet('Why They Lose')
wr.append(['LOSS REASONS (count)']); wr['A1'].font=Font(name=FNT,bold=True,size=12)
wr.append(['Reason','Count'])
for c in range(1,3): wr.cell(2,c).fill=hdr_fill; wr.cell(2,c).font=hdr_font
for _,r in lossreasons.iterrows(): wr.append([r['Loss Reason'],int(r['Count'])])
wr.column_dimensions['A'].width=42; wr.column_dimensions['B'].width=10

wb.save('model_loss_analysis.xlsx')
print('Saved model_loss_analysis.xlsx')
print('\nLOSS RATE BY NATURE:'); print(pat['Nature'].to_string())
print('\nLOSS RATE BY RULE:'); print(pat['Rule (why)'].to_string())
print('\nLOSS REASONS:'); print(lossreasons.to_string(index=False))
